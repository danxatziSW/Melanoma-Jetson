"""Sensitivity-tuned model pair across datasets, written to a formatted workbook."""
from __future__ import annotations

import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parents[3]))

import cv2
import numpy as np
import pandas as pd
import torch
import torch.nn as nn
import albumentations as A
from albumentations.pytorch import ToTensorV2
from sklearn.metrics import roc_auc_score
from torch.utils.data import DataLoader, Dataset
from tqdm import tqdm
from openpyxl.styles import PatternFill, Font, Alignment
from openpyxl.utils import get_column_letter
import openpyxl

from src.models.registry import build_model, uses_metadata
from src.utils.config import load_config
from src.utils.io import resolve_dataset_paths

MODEL_1  = "resnet50"
TRAIN_1  = "isic2019"
MODEL_2  = "efficientnet_b2"
TRAIN_2  = "isic2020"
AUG_MODE = "none"
SENS_ID  = "none_sens"   # from train_sensitivity_pair.py

ALL_DATASETS = ["ham10000", "isic2019", "isic2020"]

# fine-tuned models output higher probs, so the sweep runs past 0.65
THRESHOLDS = np.round(np.arange(0.20, 0.86, 0.005), 3)

_MEAN      = (0.485, 0.456, 0.406)
_STD       = (0.229, 0.224, 0.225)
_SITE_CATS = [
    "head/neck", "upper extremity", "lower extremity",
    "torso", "palms/soles", "oral/genital",
]


class EvalDataset(Dataset):
    def __init__(self, df: pd.DataFrame, input_size: int, with_meta: bool = False):
        self.df        = df.reset_index(drop=True)
        self.with_meta = with_meta
        self.transform = A.Compose([
            A.Resize(height=int(input_size * 1.1), width=int(input_size * 1.1)),
            A.CenterCrop(height=input_size, width=input_size),
            A.Normalize(mean=_MEAN, std=_STD),
            ToTensorV2(),
        ])
        if with_meta:
            self._encode_metadata()

    def _encode_metadata(self) -> None:
        df      = self.df
        age_col = "age_approx" if "age_approx" in df.columns else None
        self.age = (
            (df[age_col].fillna(df[age_col].median()) / 100.0).values.astype(np.float32)
            if age_col else np.zeros(len(df), dtype=np.float32)
        )
        sex_col = "sex" if "sex" in df.columns else None
        self.sex = (
            df[sex_col].map({"male": 1.0, "female": 0.0}).fillna(0.5).values.astype(np.float32)
            if sex_col else np.full(len(df), 0.5, dtype=np.float32)
        )
        site_col = "anatom_site_general_challenge" if "anatom_site_general_challenge" in df.columns else None
        self.site_ohe = np.zeros((len(df), len(_SITE_CATS)), dtype=np.float32)
        if site_col:
            site_s = df[site_col].fillna("unknown")
            for i, cat in enumerate(_SITE_CATS):
                self.site_ohe[:, i] = (site_s == cat).astype(np.float32)

    def __len__(self) -> int:
        return len(self.df)

    def __getitem__(self, idx: int):
        row   = self.df.iloc[idx]
        image = cv2.imread(str(row["image_path"]))
        image = (
            cv2.cvtColor(image, cv2.COLOR_BGR2RGB)
            if image is not None
            else np.zeros((224, 224, 3), dtype=np.uint8)
        )
        img_t = self.transform(image=image)["image"]
        label = int(row["binary_label"])
        if self.with_meta:
            meta = np.concatenate([[self.age[idx], self.sex[idx]], self.site_ohe[idx]])
            return img_t, torch.from_numpy(meta), label
        return img_t, label


def _load_test_df(splits_dir: Path, dataset_source: str, melanoma_root: Path) -> pd.DataFrame:
    df = pd.read_csv(splits_dir / "cls_test.csv")
    df = df[df["dataset_source"] == dataset_source].copy()
    df["binary_label"] = (df["label_str"] == "mel").astype(int)
    df = resolve_dataset_paths(df, melanoma_root)
    return df.reset_index(drop=True)


def _load_model(model_name: str, train_dataset: str,
                device: torch.device, noseg_dir: Path) -> tuple[nn.Module, object]:
    run_id    = f"{model_name}_{SENS_ID}"
    ckpt_path = noseg_dir / train_dataset / run_id / "checkpoints" / f"{run_id}.pt"
    if not ckpt_path.exists():
        raise FileNotFoundError(
            f"Sensitivity checkpoint not found: {ckpt_path}\n"
            "Run scripts/no_seg/train_sensitivity_pair.py first."
        )
    cfg   = load_config(model_name)
    model = build_model(model_name, cfg, num_classes=2)
    model.load_state_dict(torch.load(ckpt_path, map_location=device, weights_only=True))
    model.to(device).eval()
    print(f"  Loaded  {run_id}  ({ckpt_path})")
    return model, cfg


def _run_inference(model: nn.Module, loader: DataLoader,
                   device: torch.device, with_meta: bool) -> tuple[np.ndarray, np.ndarray]:
    all_probs, all_labels = [], []
    with torch.no_grad():
        for batch in tqdm(loader, desc="    infer", leave=False,
                          unit="batch", dynamic_ncols=True, file=sys.stdout):
            if with_meta:
                imgs, mdata, labels = batch
                logits = model(imgs.to(device), mdata.to(device))
            else:
                imgs, labels = batch
                logits = model(imgs.to(device))
            probs = torch.softmax(logits, dim=1)[:, 1].cpu().numpy()
            all_probs.extend(probs)
            all_labels.extend(labels.numpy())
    return np.array(all_probs), np.array(all_labels)


def _metrics(probs: np.ndarray, labels: np.ndarray, thr: float) -> dict:
    preds = (probs >= thr).astype(int)
    tp = int(((preds == 1) & (labels == 1)).sum())
    tn = int(((preds == 0) & (labels == 0)).sum())
    fp = int(((preds == 1) & (labels == 0)).sum())
    fn = int(((preds == 0) & (labels == 1)).sum())
    sens = tp / max(tp + fn, 1)
    spec = tn / max(tn + fp, 1)
    prec = tp / max(tp + fp, 1)
    f1   = (2 * prec * sens) / max(prec + sens, 1e-9)
    f2   = (5 * prec * sens) / max(4 * prec + sens, 1e-9)
    acc  = (tp + tn) / max(tp + tn + fp + fn, 1)
    return dict(
        tp=tp, tn=tn, fp=fp, fn=fn,
        sensitivity=round(sens, 4), specificity=round(spec, 4),
        precision=round(prec, 4),   accuracy=round(acc, 4),
        f1=round(f1, 4),            f2=round(f2, 4),
    )


_HDR  = PatternFill(fill_type="solid", fgColor="1F4E79")
_GRN  = PatternFill(fill_type="solid", fgColor="E2EFDA")   # best F1
_YEL  = PatternFill(fill_type="solid", fgColor="FFF2CC")   # best F2
_ORG  = PatternFill(fill_type="solid", fgColor="FCE4D6")   # sens >= 0.80
_BLU  = PatternFill(fill_type="solid", fgColor="DDEBF7")   # thr=0.50
_EVEN = PatternFill(fill_type="solid", fgColor="F2F2F2")
_HFNT = Font(bold=True, color="FFFFFF", size=10)
_BFNT = Font(bold=True, size=10)


def _style_ws(ws, highlights: dict[int, PatternFill]) -> None:
    for cell in ws[1]:
        cell.fill      = _HDR
        cell.font      = _HFNT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 28
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
        fill = highlights.get(row_idx, _EVEN if row_idx % 2 == 0 else None)
        for cell in row:
            if fill:
                cell.fill = fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            if isinstance(cell.value, float):
                cell.number_format = "0.0000"
    for col_idx in range(1, ws.max_column + 1):
        letter = get_column_letter(col_idx)
        val = ws.cell(row=1, column=col_idx).value or ""
        ws.column_dimensions[letter].width = max(12, len(str(val)) + 4)
    ws.freeze_panes = "A2"


def main() -> None:
    base_cfg   = load_config()
    device     = torch.device("cuda" if torch.cuda.is_available() else "cpu")
    splits_dir = Path(base_cfg.paths.data_splits)
    melanoma_root = Path(base_cfg.paths.melanoma_data)
    noseg_dir  = Path(base_cfg.paths.outputs) / "ablation_noseg"
    out_xlsx   = noseg_dir / "evaluation_sens_pair.xlsx"

    pair_name = f"{MODEL_1}_{TRAIN_1}_{SENS_ID}+{MODEL_2}_{TRAIN_2}_{SENS_ID}"
    gpu_label = torch.cuda.get_device_name(0) if device.type == "cuda" else "CPU"

    print(f"\n{'='*72}")
    print(f"  Sensitivity pair evaluation (no-seg)")
    print(f"  Pair   : {MODEL_1}/{TRAIN_1} + {MODEL_2}/{TRAIN_2}  (suffix: _{SENS_ID})")
    print(f"  Device : {gpu_label}")
    print(f"  Sweep  : {THRESHOLDS[0]:.3f} → {THRESHOLDS[-1]:.3f}  "
          f"(step 0.005, {len(THRESHOLDS)} points, wider to handle shifted probs)")
    print(f"{'='*72}\n")

    test_dfs: dict[str, pd.DataFrame] = {}
    for ds in ALL_DATASETS:
        df    = _load_test_df(splits_dir, ds, melanoma_root)
        n_mel = int(df["binary_label"].sum())
        test_dfs[ds] = df
        print(f"  {ds.upper():<12} total={len(df)}  mel={n_mel}  non-mel={len(df)-n_mel}")
    print()

    model1, cfg1 = _load_model(MODEL_1, TRAIN_1, device, noseg_dir)
    model2, cfg2 = _load_model(MODEL_2, TRAIN_2, device, noseg_dir)
    print()

    ds_probs: dict[str, tuple[np.ndarray, np.ndarray]] = {}
    ds_aucs:  dict[str, float] = {}

    for ds in ALL_DATASETS:
        print(f"  Inferring on {ds.upper()} ...")

        meta1   = uses_metadata(MODEL_1)
        loader1 = DataLoader(
            EvalDataset(test_dfs[ds], getattr(cfg1, "input_size", 224), with_meta=meta1),
            batch_size=cfg1.batch_size, shuffle=False,
            num_workers=getattr(cfg1, "num_workers", 4),
        )
        p1, labels = _run_inference(model1, loader1, device, meta1)

        meta2   = uses_metadata(MODEL_2)
        loader2 = DataLoader(
            EvalDataset(test_dfs[ds], getattr(cfg2, "input_size", 224), with_meta=meta2),
            batch_size=cfg2.batch_size, shuffle=False,
            num_workers=getattr(cfg2, "num_workers", 4),
        )
        p2, _ = _run_inference(model2, loader2, device, meta2)

        mean_p = (p1 + p2) / 2.0
        try:
            auc = float(roc_auc_score(labels, mean_p))
        except Exception:
            auc = float("nan")
        ds_probs[ds] = (mean_p, labels)
        ds_aucs[ds]  = auc
        print(f"    AUC = {auc:.4f}  "
              f"(prob range: {mean_p.min():.3f} to {mean_p.max():.3f}  "
              f"median: {float(np.median(mean_p)):.3f})")

    del model1, model2
    if device.type == "cuda":
        torch.cuda.empty_cache()

    print(f"\n  Sweeping {len(THRESHOLDS)} thresholds ...")
    sweep_rows = []
    for thr in THRESHOLDS:
        per_ds: dict[str, dict] = {ds: _metrics(p, l, thr) for ds, (p, l) in ds_probs.items()}
        avg_f1   = float(np.mean([m["f1"]          for m in per_ds.values()]))
        avg_f2   = float(np.mean([m["f2"]          for m in per_ds.values()]))
        avg_sens = float(np.mean([m["sensitivity"] for m in per_ds.values()]))
        avg_spec = float(np.mean([m["specificity"] for m in per_ds.values()]))
        avg_acc  = float(np.mean([m["accuracy"]    for m in per_ds.values()]))
        row: dict = dict(
            threshold=round(thr, 3),
            avg_f1=round(avg_f1, 4),   avg_f2=round(avg_f2, 4),
            avg_sensitivity=round(avg_sens, 4), avg_specificity=round(avg_spec, 4),
            avg_accuracy=round(avg_acc, 4),
        )
        for ds, m in per_ds.items():
            row[f"f1_{ds}"]          = m["f1"]
            row[f"f2_{ds}"]          = m["f2"]
            row[f"sensitivity_{ds}"] = m["sensitivity"]
            row[f"specificity_{ds}"] = m["specificity"]
            row[f"accuracy_{ds}"]    = m["accuracy"]
            row[f"tp_{ds}"]          = m["tp"]
            row[f"tn_{ds}"]          = m["tn"]
            row[f"fp_{ds}"]          = m["fp"]
            row[f"fn_{ds}"]          = m["fn"]
            row[f"auc_{ds}"]         = round(ds_aucs[ds], 4)
        sweep_rows.append(row)

    sweep_df = pd.DataFrame(sweep_rows)

    f1_idx  = int(sweep_df["avg_f1"].idxmax())
    f2_idx  = int(sweep_df["avg_f2"].idxmax())
    t50_idx = int((sweep_df["threshold"] - 0.50).abs().idxmin())
    s80_mask = sweep_df["avg_sensitivity"] >= 0.80
    s80_idx  = int(sweep_df[s80_mask].index[0]) if s80_mask.any() else None

    r_f1  = sweep_df.iloc[f1_idx]
    r_f2  = sweep_df.iloc[f2_idx]
    r_t50 = sweep_df.iloc[t50_idx]

    sep = "=" * 72
    print(f"\n{sep}")
    print(f"  Pair   : {pair_name}")
    print(f"  AUCs   : " + "  ".join(f"{ds.upper()}={ds_aucs[ds]:.4f}" for ds in ALL_DATASETS))

    print(f"\n  ┌──────────────────┬────────┬────────┬────────┬────────┬────────┐")
    print(f"  │  Criterion       │  Thr   │ Avg F1 │ Avg F2 │Avg Sns │Avg Spc │")
    print(f"  ├──────────────────┼────────┼────────┼────────┼────────┼────────┤")
    print(f"  │  Best F1         │ {r_f1['threshold']:.3f}  │ {r_f1['avg_f1']:.4f} │ {r_f1['avg_f2']:.4f} │ {r_f1['avg_sensitivity']:.4f} │ {r_f1['avg_specificity']:.4f} │")
    print(f"  │  Best F2 (β=2)   │ {r_f2['threshold']:.3f}  │ {r_f2['avg_f1']:.4f} │ {r_f2['avg_f2']:.4f} │ {r_f2['avg_sensitivity']:.4f} │ {r_f2['avg_specificity']:.4f} │")
    if s80_idx is not None:
        r_s80 = sweep_df.iloc[s80_idx]
        print(f"  │  Sens ≥ 0.80     │ {r_s80['threshold']:.3f}  │ {r_s80['avg_f1']:.4f} │ {r_s80['avg_f2']:.4f} │ {r_s80['avg_sensitivity']:.4f} │ {r_s80['avg_specificity']:.4f} │")
    else:
        print(f"  │  Sens ≥ 0.80     │  N/A   │   N/A  │   N/A  │   N/A  │   N/A  │")
    print(f"  │  Thr = 0.50      │ {r_t50['threshold']:.3f}  │ {r_t50['avg_f1']:.4f} │ {r_t50['avg_f2']:.4f} │ {r_t50['avg_sensitivity']:.4f} │ {r_t50['avg_specificity']:.4f} │")
    print(f"  └──────────────────┴────────┴────────┴────────┴────────┴────────┘")

    print(f"\n  Per-dataset at Best F1 (thr={r_f1['threshold']:.3f}):")
    for ds in ALL_DATASETS:
        print(f"    {ds.upper():<12}  AUC={ds_aucs[ds]:.4f}  "
              f"F1={r_f1[f'f1_{ds}']:.4f}  F2={r_f1[f'f2_{ds}']:.4f}  "
              f"sens={r_f1[f'sensitivity_{ds}']:.4f}  spec={r_f1[f'specificity_{ds}']:.4f}  "
              f"acc={r_f1[f'accuracy_{ds}']:.4f}  "
              f"TP={int(r_f1[f'tp_{ds}'])}  FN={int(r_f1[f'fn_{ds}'])}")

    print(f"\n  Per-dataset at Best F2 (thr={r_f2['threshold']:.3f}):")
    for ds in ALL_DATASETS:
        print(f"    {ds.upper():<12}  AUC={ds_aucs[ds]:.4f}  "
              f"F1={r_f2[f'f1_{ds}']:.4f}  F2={r_f2[f'f2_{ds}']:.4f}  "
              f"sens={r_f2[f'sensitivity_{ds}']:.4f}  spec={r_f2[f'specificity_{ds}']:.4f}  "
              f"acc={r_f2[f'accuracy_{ds}']:.4f}  "
              f"TP={int(r_f2[f'tp_{ds}'])}  FN={int(r_f2[f'fn_{ds}'])}")

    print(f"\n  Sweep (every 0.05 + key thresholds):")
    print(f"  {'Thr':>6}  {'Avg F1':>8}  {'Avg F2':>8}  {'Avg Sens':>10}  {'Avg Spec':>10}  {'Avg Acc':>9}")
    key_thrs = {r_f1["threshold"], r_f2["threshold"], 0.50}
    if s80_idx is not None:
        key_thrs.add(sweep_df.iloc[s80_idx]["threshold"])
    for _, r in sweep_df.iterrows():
        if r["threshold"] % 0.05 < 0.003 or r["threshold"] in key_thrs:
            markers = ""
            if abs(r["threshold"] - r_f1["threshold"]) < 0.001: markers += " ★F1"
            if abs(r["threshold"] - r_f2["threshold"]) < 0.001: markers += " ★F2"
            if s80_idx and abs(r["threshold"] - sweep_df.iloc[s80_idx]["threshold"]) < 0.001: markers += " ★S80"
            print(f"  {r['threshold']:>6.3f}  {r['avg_f1']:>8.4f}  {r['avg_f2']:>8.4f}  "
                  f"{r['avg_sensitivity']:>10.4f}  {r['avg_specificity']:>10.4f}  "
                  f"{r['avg_accuracy']:>9.4f}{markers}")

    print(f"\n  Saved: {out_xlsx}\n{sep}\n")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    sweep_cols = ["threshold", "avg_f1", "avg_f2", "avg_sensitivity",
                  "avg_specificity", "avg_accuracy"]
    for ds in ALL_DATASETS:
        sweep_cols += [f"f1_{ds}", f"f2_{ds}", f"sensitivity_{ds}",
                       f"specificity_{ds}", f"accuracy_{ds}",
                       f"tp_{ds}", f"tn_{ds}", f"fp_{ds}", f"fn_{ds}", f"auc_{ds}"]
    sweep_cols = [c for c in sweep_cols if c in sweep_df.columns]

    ws_sweep = wb.create_sheet("Threshold_Sweep")
    ws_sweep.append(sweep_cols)
    hl_sweep: dict[int, PatternFill] = {}
    for i, (_, r) in enumerate(sweep_df[sweep_cols].iterrows(), start=2):
        ws_sweep.append(list(r))
        t = r["threshold"]
        if abs(t - r_f1["threshold"])  < 0.001: hl_sweep[i] = _GRN
        elif abs(t - r_f2["threshold"]) < 0.001: hl_sweep[i] = _YEL
        elif s80_idx and abs(t - sweep_df.iloc[s80_idx]["threshold"]) < 0.001: hl_sweep[i] = _ORG
        elif abs(t - 0.50) < 0.001: hl_sweep[i] = _BLU
    _style_ws(ws_sweep, hl_sweep)

    ws_key = wb.create_sheet("Key_Thresholds")
    key_cols = ["criterion", "threshold", "avg_f1", "avg_f2",
                "avg_sensitivity", "avg_specificity", "avg_accuracy"]
    for ds in ALL_DATASETS:
        key_cols += [f"f1_{ds}", f"f2_{ds}", f"sensitivity_{ds}",
                     f"specificity_{ds}", f"accuracy_{ds}", f"tp_{ds}", f"fn_{ds}"]
    key_cols = [c for c in key_cols if c in sweep_df.columns or c == "criterion"]

    key_entries = [("Best F1", r_f1, _GRN), ("Best F2 (β=2)", r_f2, _YEL),
                   ("Thr = 0.50", r_t50, _BLU)]
    if s80_idx is not None:
        key_entries.append(("Sens ≥ 0.80", sweep_df.iloc[s80_idx], _ORG))

    ws_key.append(key_cols)
    hl_key: dict[int, PatternFill] = {}
    for i, (label, row, fill) in enumerate(key_entries, start=2):
        vals = {"criterion": label}
        vals.update({c: row[c] for c in key_cols if c != "criterion" and c in row.index})
        ws_key.append([vals.get(c, "") for c in key_cols])
        hl_key[i] = fill
    _style_ws(ws_key, hl_key)

    r_start = len(key_entries) + 3
    ws_key.cell(row=r_start, column=1).value = "Legend"
    ws_key.cell(row=r_start, column=1).font  = _BFNT
    for i, (col, desc) in enumerate([
        ("Green",  "Best F1: maximises avg F1 across 3 datasets"),
        ("Yellow", "Best F2: maximises avg F2 (beta=2), weights sensitivity 2x"),
        ("Orange", "Sens >= 0.80: lowest threshold reaching 80% avg sensitivity"),
        ("Blue",   "Thr = 0.50: standard baseline"),
    ], start=r_start + 1):
        ws_key.cell(row=i, column=1).value = col
        ws_key.cell(row=i, column=2).value = desc

    for ds in ALL_DATASETS:
        if f"f1_{ds}" not in sweep_df.columns:
            continue
        ds_cols = ["threshold"] + [c for c in sweep_df.columns if c.endswith(f"_{ds}")]
        ws_ds = wb.create_sheet(ds.upper())
        ws_ds.append(ds_cols)
        hl_ds: dict[int, PatternFill] = {}
        for i, (_, r) in enumerate(sweep_df[ds_cols].iterrows(), start=2):
            ws_ds.append(list(r))
            t = sweep_df.iloc[i - 2]["threshold"]
            if abs(t - r_f1["threshold"])  < 0.001: hl_ds[i] = _GRN
            elif abs(t - r_f2["threshold"]) < 0.001: hl_ds[i] = _YEL
            elif s80_idx and abs(t - sweep_df.iloc[s80_idx]["threshold"]) < 0.001: hl_ds[i] = _ORG
        _style_ws(ws_ds, hl_ds)

    wb.save(out_xlsx)
    print(f"  Excel saved: {out_xlsx}")


if __name__ == "__main__":
    main()
